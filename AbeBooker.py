from openpyxl import load_workbook
from urllib.request import urlopen, Request
import urllib.error
import time
import datetime
import re
from random import randint
##Openpyxl requires it's own installation via 'pip install openpyxl'
workbook = load_workbook(filename="books.xlsx")
totalPrice = 0.00
httpErrorCount = 0 
indexErrorCount = 0
rowCount = 1
workbook.sheetnames
sheet = workbook.active
rownum = input("How many books do you have?\n")
rownum = int(rownum)
#colnum = input("Which number column has your ISBN's?\(i.e A = 1, B = 2, C = 3 \)\n")
#colnum = int(colnum)
#cell_range = sheet['C2':'C'+colnum]
#print(cell_range) # returns a list of the non null colomns in row 1
#Iterates through the workbook at 
#for row in sheet.iter_rows(min_row=rownum, min_col =colnum, max_col=colnum, max_row=int(rownum), values_only=True):

for row in sheet.iter_rows(min_row=2, min_col =3, max_col=3, max_row=int(rownum), values_only=True):

	ISBNV = row[0]
	currentTime = datetime.datetime.now()
	#Handle "N/A" values in ISBN here. Maybe a switchcase? 
	try:
		url = "https://www.abebooks.com/servlet/SearchResults?kn=" + ISBNV
		print(url)
	except TypeError:
		print("No value at row " + str(rowCount) + "\n")
	try:
		page = urlopen(Request(url, headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36'}))# Maybe look into creating an array of acceptable user agents that gets incremented on errors? 
		html_bytes = page.read()
		html = html_bytes.decode("latin-1")
		#html = html_bytes.decode("utf-8")
		listedPrices = re.findall("US\$\s(\d+\.\d+)", html)
		try: # This handles the regex if a book doesn't have a price on the page we search up. This error can happen because some books don't have ISBN's
			price = listedPrices[0]
			totalPrice += float(price)
			print(ISBNV + " costs $"+ price + " requested at " + str(currentTime))
			print("At row " + str(rowCount) + " Total price is bout $"+ str(totalPrice)+"\n")
			time.sleep(randint(10,30)) #20 seconds is too quick, 
			rowCount+= 1
		except IndexError:
			print(ISBNV + " Is causing an error at row "+ str(rowCount) + " at "+ str(currentTime)+ "!\n")
			#Create an array of errored out ISBN'search
			indexErrorCount += 1
			rowCount += 1
	except ConnectionResetError:
		print("Closed Connection. Waiting. Skipping " + str(ISBNV) + "at row " + str(rowCount)+ "Time: "+ str(currentTime)+"\n")
	except urllib.error.URLError as e:
		print(e.__dict__)
		print("\nError 429 at "+ str(currentTime)+ "\n")
		time.sleep(randint(10,15))
		#Implement error counter
		httpErrorCount += 1
	

	
print("Everything all together's bout $"+str(totalPrice))
print(str(httpErrorCount) + " HTTP Errors\n")
print(str(indexErrorCount) + "Index Errors\n")
